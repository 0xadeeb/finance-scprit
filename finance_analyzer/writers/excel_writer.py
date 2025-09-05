"""
Excel file writing and formatting functionality.
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from typing import Dict

from .base import Writer
from ..constants import CREDIT_CATEGORIES, DEBIT_CATEGORIES


class ExcelWriter(Writer):
    """Handles Excel file writing and formatting"""
    
    def __init__(self):
        self.blue_font = Font(color="0000FF")
        self.red_font = Font(color="FF0000")
        self.maroon_font = Font(color="800000")
        
        # Define fill colors for different months
        self.fill_array = [
            PatternFill(start_color='d0e8ca', end_color='d0e8ca', fill_type='solid'),  # Green
            PatternFill(start_color='e8caca', end_color='e8caca', fill_type='solid'),  # Red
            PatternFill(start_color='d5e1f5', end_color='d5e1f5', fill_type='solid'),  # Blue
            PatternFill(start_color='f0e0b4', end_color='f0e0b4', fill_type='solid')   # Yellow
        ]
        
        # Define border style
        self.thin_border = Border(
            left=Side(style='dotted'),
            right=Side(style='dotted'),
            top=Side(style='dotted'),
            bottom=Side(style='dotted')
        )
    
    def append_to_sheet(self, month: int, workbook: openpyxl.Workbook, sheet_name: str, 
                       dataframe: pd.DataFrame) -> tuple:
        """Append data to a worksheet with formatting"""
        # If the sheet does not exist, create it
        if sheet_name not in workbook.sheetnames:
            workbook.create_sheet(sheet_name)
        worksheet = workbook[sheet_name]

        start_row = worksheet.max_row + 1
        fill_colour = self.fill_array[month % len(self.fill_array)]

        for r_idx, row in enumerate(dataframe.itertuples(index=False), start=start_row):
            for c_idx, value in enumerate(row, start=1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                cell.fill = fill_colour
                cell.border = self.thin_border

        return worksheet, start_row
    
    def get_file_extension(self) -> str:
        """Get the file extension for this writer"""
        return ".xlsx"
    
    def get_writer_name(self) -> str:
        """Get the name of this writer"""
        return "Excel Writer"
    
    def supports_formatting(self) -> bool:
        """Check if this writer supports advanced formatting"""
        return True
    
    def format_summary_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                                summary_df: pd.DataFrame) -> None:
        """Apply formatting to the summary worksheet"""
        credit_start_row = 4
        debit_start_row = credit_start_row + len(CREDIT_CATEGORIES) + 1
        net_start_row = debit_start_row + len(DEBIT_CATEGORIES) + 1

        # Apply font color to the rows based on whether they are credit or debit rows
        for row in range(credit_start_row, worksheet.max_row + 1):
            cell = worksheet.cell(row=row, column=1)
            if credit_start_row <= row < debit_start_row:
                cell.font = self.blue_font
            elif debit_start_row <= row < net_start_row:
                cell.font = self.red_font
            elif net_start_row <= row < len(summary_df):
                cell.font = self.maroon_font

        # Adjust width of columns
        index_max_length = max(len(str(i)) for i in summary_df.index)
        worksheet.column_dimensions['A'].width = index_max_length + 2

        for i, col in enumerate(summary_df.columns):
            column_len = summary_df[col].astype(str).map(len).max()
            column_len = max(column_len, len(col)) + 2
            worksheet.column_dimensions[get_column_letter(i + 2)].width = column_len
    
    def format_transactions_worksheet(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                    transactions_df: pd.DataFrame, start_row: int) -> None:
        """Apply formatting to transactions worksheet"""
        # Apply font colors based on credit/debit
        for i in range(len(transactions_df)):
            if pd.notna(transactions_df.iloc[i]['Credit']):
                worksheet.cell(row=i + start_row, column=5).font = self.blue_font
            else:
                worksheet.cell(row=i + start_row, column=5).font = self.red_font

        # Adjust column widths
        for i, col in enumerate(transactions_df.columns):
            column_len = transactions_df[col].astype(str).map(len).max()
            column_len = max(column_len, len(col)) + 2
            worksheet.column_dimensions[get_column_letter(i + 1)].width = column_len
    
    def write(self, year: str, month: int, transactions_dfs: Dict[str, pd.DataFrame], 
             summary_df: pd.DataFrame, output_path: str) -> bool:
        """Write all data to Excel file with formatting"""
        try:
            self.write_to_excel(year, month, transactions_dfs, summary_df, output_path)
            return True
        except Exception as e:
            print(f"❌ Error writing Excel file: {str(e)}")
            return False
    
    def write_to_excel(self, year: str, month: int, transactions_dfs: Dict[str, pd.DataFrame], 
                      summary_df: pd.DataFrame, output_path: str) -> None:
        """Write all data to Excel file with formatting"""
        
        # Write summary data first
        with pd.ExcelWriter(output_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            summary_df.to_excel(writer, sheet_name=year)
            summary_worksheet = writer.sheets[year]
            self.format_summary_worksheet(summary_worksheet, summary_df)
        
        # Brief pause to ensure file is fully written and closed
        import time
        time.sleep(0.1)

        # Write transaction data
        workbook = openpyxl.load_workbook(output_path)
        try:
            for bank_name, transactions_df in transactions_dfs.items():
                transactions_worksheet, start_row = self.append_to_sheet(month, workbook, bank_name, transactions_df)
                self.format_transactions_worksheet(transactions_worksheet, transactions_df, start_row)

            workbook.save(output_path)
        finally:
            # Ensure workbook is properly closed
            workbook.close()
