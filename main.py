import pandas as pd
import json
import os
import calendar

# Fow writer
import openpyxl
from openpyxl.styles import Font, PatternFill, Side, Border
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import FormulaRule

class bank:
    def __init__(self, name, skiprows, date, description, debit, credit, parse_fn):
        self.name = name
        self.skiprows = skiprows
        self.date = date
        self.description = description
        self.credit = credit
        self.debit = debit
        self.parse_fn = parse_fn


INPUT_DIR = '/content/drive/MyDrive/Bank/'
OUTPUT_FILE = '/content/drive/MyDrive/Bank/Finance.xlsx'
MERCHANT_CATEGORY_FILE = '/content/drive/MyDrive/Bank/merchant_category.json'
BALANCE_ROWS = ['Opening Bank Bal.', 'Opening In Hand', 'Closing In Hand', 'Closing Bank Bal.']

CREDIT_CATEGORIES = [
    "Salary",
]

DEBIT_CATEGORIES = [
    "Food",
    "House",                    # House rent, furnishing etc
    "Entertainment",
    "Clothes & Accessories",    # Cloths and accessories
    "Transportation",           # Cabs, petrol, cleaning vehicles etc
    "Health",                   # Doctor, Medicines etc
    "Fitness",                 # Gym, football, boxing etc
    "Grooming",
    "Grocery",                  # Household items - Utensils, cleaning products etc
    "Network",
    "Free",                     # Give money for free to friends or kins
    "Misl"                      # General category or don't know the category
]
NET_CATEGORIES = [
    "Investment",
    "Deposit",
    "Lend",
    "Borrow"
] # Non Expenditure categories - not being used now

CATEGORIES = CREDIT_CATEGORIES + DEBIT_CATEGORIES + NET_CATEGORIES

CAT_MAPPING = {
    "cloths": "Clothes & Accessories",
    "accessories": "Clothes & Accessories",
    "petrol": "Transportation",
    "travel": "Transportation",
    "transport": "Transportation",
    "cab": "Transportation",
    "utilities": "Grocery",
    "doctor": "Health",
    "rent": "House",
    "medicine": "Health",
    "other": "Misl"
}

BANK_ACCOUNTS = ["SBI", "HDFC"]

def parse_sbi(transaction_details: pd.DataFrame) -> tuple[str, str]:
    description = transaction_details['Description'].lower()
    words = description.split('-')
    transaction_type = words[0]
    category = None
    merchant = None

    if len(words) > 1 and 'upi' in words[1]:
        words = words[1].split('/')
        transaction_type = words[0]
        if transaction_type == "upi":
            category = words[-1]
            merchant = words[3]

    return category, merchant

def parse_hdfc(transaction_details: pd.DataFrame) -> tuple[str, str]:
    description = transaction_details['Description'].lower()
    words = description.split('-')
    transaction_type = words[0].lower()
    category = None
    merchant = None

    if transaction_type == 'upi':
        category = words[-1]
        merchant = words[2]

    if 'neft' in transaction_type:
        category = words[-1]
        merchant = words[2]

    return category, merchant


BANK_DETAILS = {
    'hdfc': bank('HDFC', 20, 'Date', 'Narration', 'Withdrawal Amt.', 'Deposit Amt.', parse_hdfc),
    'sbi': bank('SBI', 19, 'Txn Date', 'Description', 'Debit', 'Credit', parse_sbi)
}

def prompt_user_for_category(transaction_details: pd.DataFrame) -> tuple[str, bool]:
    #return "Misl", False
    print(f"Transaction details:")
    print(f"Date: {transaction_details['Date']}")
    print(f"Description: {transaction_details['Description']}")
    print(f"Amount: {transaction_details['Amount']}")
    print("Please select a category:")

    num_categories = len(CATEGORIES)
    for i in range(0, num_categories, 4):
        row = []
        for j in range(i, min(i + 4, num_categories)):
            row.append(f"{j+1}. {CATEGORIES[j].title()}")
        print("{:<25} {:<25} {:<25} {:<25}".format(*row, *([""]*(4-len(row)))))

    while True:
        try:
            selected_index = int(input("\nEnter the number corresponding to the category: ")) - 1
            if selected_index < 0 or selected_index >= num_categories:
                print("Invalid selection. Please select a number corresponding to the category.")
                continue
            break
        except ValueError:
            print("Invalid input. Please enter a number.")

    while True:
        store_mapping = input("Do you want to remember this merchant-category mapping? (y/[n]): ").lower()
        if store_mapping not in ['y', 'n', '']:
            print("Invalid input. Please enter 'y' or 'n'.")
        else:
            store_mapping = store_mapping == 'y'
            break

    print()
    return CATEGORIES[selected_index], store_mapping

def append_to_sheet(month: int, workbook: any, sheet_name: str, dataframe: pd.DataFrame) -> tuple[any, int]:
    # Define a light green fill
    fill_array = [
        PatternFill(start_color='d0e8ca', end_color='d0e8ca', fill_type='solid'), # Green
        PatternFill(start_color='e8caca', end_color='e8caca', fill_type='solid'), # Red
        PatternFill(start_color='d5e1f5', end_color='d5e1f5', fill_type='solid'), # Blue
        PatternFill(start_color='f0e0b4', end_color='f0e0b4', fill_type='solid')  # Yellow
    ]

    # Define a thin border style
    thin_border = Border(
        left=Side(style='dotted'),
        right=Side(style='dotted'),
        top=Side(style='dotted'),
        bottom=Side(style='dotted')
    )

    # If the sheet does not exist, create it
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)
    worksheet = workbook[sheet_name]

    start_row = worksheet.max_row + 1
    fill_colour = fill_array[month % len(fill_array)]

    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=1):
            cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
            cell.fill = fill_colour
            cell.border = thin_border

    return worksheet, start_row

def write_to_excel(year: str, month: int, transactions_dfs: list[pd.DataFrame], summary_df: pd.DataFrame) -> None:
    blue_font = Font(color="0000FF")
    red_font = Font(color="FF0000")
    maroon_font = Font(color="800000")

    workbook = openpyxl.load_workbook(OUTPUT_FILE)

    with pd.ExcelWriter(OUTPUT_FILE, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        summary_df.to_excel(writer, sheet_name=year)
        summary_worksheet = writer.sheets[year]

        credit_start_row = 4
        debit_start_row = credit_start_row + len(CREDIT_CATEGORIES) + 1
        net_start_row = debit_start_row + len(DEBIT_CATEGORIES) + 1

        # Apply font color to the rows based on whether they are credit or debit rows
        for row in range(credit_start_row, summary_worksheet.max_row + 1):
            cell = summary_worksheet.cell(row=row, column=1)  # Assuming the credits/debits are in the first column
            if credit_start_row <= row < debit_start_row:
                cell.font = blue_font
            elif debit_start_row <= row < net_start_row:
                cell.font = red_font
            elif net_start_row <= row < len(summary_df):
                cell.font = maroon_font

        # Adjust width of columns
        index_max_length = max(len(str(i)) for i in summary_df.index)
        summary_worksheet.column_dimensions['A'].width = index_max_length + 2

        for i, col in enumerate(summary_df.columns):
            column_len = summary_df[col].astype(str).map(len).max()
            column_len = max(column_len, len(col)) + 2
            summary_worksheet.column_dimensions[get_column_letter(i + 2)].width = column_len

    workbook = openpyxl.load_workbook(OUTPUT_FILE)
    for bank_name, transactions_df in transactions_dfs.items():
        transactions_worksheet, start_row = append_to_sheet(month, workbook, bank_name, transactions_df)
        for i in range(len(transactions_df)):
            if pd.notna(transactions_df.loc[i, 'Credit']):
                transactions_worksheet.cell(row= i + start_row, column=5).font = blue_font
            else:
                transactions_worksheet.cell(row= i + start_row, column=5).font = red_font

        for i, col in enumerate(transactions_df.columns):
            column_len = transactions_df[col].astype(str).map(len).max()
            column_len = max(column_len, len(col)) + 2
            transactions_worksheet.column_dimensions[get_column_letter(i + 1)].width = column_len

    workbook.save(OUTPUT_FILE)

def parse(bank_name: str, transaction_details: pd.DataFrame) -> tuple[str, str]:

    parser_function = BANK_DETAILS.get(bank_name.lower()).parse_fn

    if parser_function:
        return parser_function(transaction_details)
    else:
        return None, None

def categorise_transaction(bank_name: str, transaction_details: pd.DataFrame, merchant_category_dict: dict[str, str]) -> str:
    category, merchant = parse(bank_name, transaction_details)

    if category and category in CAT_MAPPING:
        return CAT_MAPPING[category]

    if not merchant or merchant not in merchant_category_dict:
        category = None
    else:
        category = merchant_category_dict[merchant]
        if category in CATEGORIES:
            return category
        else:
            del merchant_category_dict[merchant]
            category = None

    category, store_mapping = prompt_user_for_category(transaction_details)
    if merchant and store_mapping:
        merchant_category_dict[merchant] = category

    return category

def read_transactions_df(bank_name: str, year: str, month: int) -> pd.DataFrame:
    input_file = INPUT_DIR + f"{bank_name}/{year}/Bankstatement{calendar.month_abbr[month]}.xlsx"
    details = BANK_DETAILS.get(bank_name.lower())
    name_mapping = {
        details.date: 'Date',
        details.description: 'Description',
        details.credit: 'Credit',
        details.debit: 'Debit'
    }
    transactions_df = pd.read_excel(input_file, skiprows=details.skiprows)
    transactions_df['Category'] = pd.NA
    transactions_df.rename(name_mapping, axis = 'columns', inplace = True)
    transactions_df['Date'] = pd.to_datetime(transactions_df['Date'], errors='coerce', dayfirst=True)
    transactions_df['Date'] = transactions_df['Date'].dt.strftime('%d %b %Y')
    transactions_df = transactions_df[["Date", "Description", "Credit", "Debit", 'Category']]
    return transactions_df

def main() -> None:
    # year = input("Enter the year: ")
    month = int(input("Enter the month: "))
    year = "2024"
    # month = 4
    if not os.path.exists(OUTPUT_FILE):
        print("Output file does not exist")
        return

    if os.path.exists(MERCHANT_CATEGORY_FILE):
        with open(MERCHANT_CATEGORY_FILE, 'r') as f:
            merchant_category_dict = json.load(f)
    else:
        merchant_category_dict = {}

    category_sums = {category: 0 for category in CATEGORIES}

    for item in CATEGORIES:
        CAT_MAPPING[item.lower()] = item.title()

    transactions_dfs = {}

    for bank_name in BANK_ACCOUNTS:
        input_file = INPUT_DIR + f"{bank_name}/{year}/Bankstatement{calendar.month_abbr[month]}.xlsx"

        if not os.path.exists(input_file):
            print(f"Input file {input_file} doesn't exist")
            continue

        transactions_df = read_transactions_df(bank_name, year, month)

        for i, row in transactions_df.iterrows():
            amount = row['Debit'] if pd.notna(row['Debit']) else row['Credit']
            if pd.notna(row['Debit']):
                amount = -amount

            if pd.isna(amount):
                transactions_df = transactions_df.iloc[:i]
                break

            transaction_details = {
                'Date': row['Date'],
                'Description': row['Description'],
                'Amount': amount
            }

            category = categorise_transaction(bank_name, transaction_details, merchant_category_dict)
            transactions_df.loc[i, 'Category'] = category
            category_sums[category] += amount

        with open(MERCHANT_CATEGORY_FILE, 'w') as f:
            json.dump(merchant_category_dict, f)

        transactions_dfs[bank_name] = transactions_df

    with pd.ExcelFile(OUTPUT_FILE, engine='openpyxl') as xls:
        summary_df = pd.read_excel(xls, year, index_col=0)

    for category, total in category_sums.items():
        summary_df.loc[category, calendar.month_name[month]] = total

    monthly_columns = [calendar.month_name[i + 1] for i in range(12)]

    summary_df.loc['Total Credit', calendar.month_name[month]] = sum([summary_df.loc[category, calendar.month_name[month]] for category in CREDIT_CATEGORIES])
    summary_df.loc['Total Debit', calendar.month_name[month]] = sum([summary_df.loc[category, calendar.month_name[month]] for category in DEBIT_CATEGORIES])
    summary_df.loc['Total NET', calendar.month_name[month]] = sum([summary_df.loc[category, calendar.month_name[month]] for category in NET_CATEGORIES])
    summary_df['Avg'] = summary_df[monthly_columns].mean(axis=1).round(2)
    summary_df['Total'] = summary_df[monthly_columns].sum(axis=1).round(2) - summary_df['Avg']
    summary_df.loc['Closing Bank Bal.', calendar.month_name[month]] = summary_df.loc['Opening Bank Bal.', calendar.month_name[month]] + summary_df.loc['Total Credit', calendar.month_name[month]] + summary_df.loc['Total Debit', calendar.month_name[month]]
    summary_df.loc['Closing In Hand', calendar.month_name[month]] = summary_df.loc['Opening In Hand', calendar.month_name[month]]

    if month != 12: # Need to handle next year
        summary_df.loc['Opening Bank Bal.', calendar.month_name[month + 1]] = summary_df.loc['Closing Bank Bal.', calendar.month_name[month]]
        summary_df.loc['Opening In Hand', calendar.month_name[month + 1]] = summary_df.loc['Closing In Hand', calendar.month_name[month]]


    ordered_rows = BALANCE_ROWS[:2] + CREDIT_CATEGORIES + ["Total Credit"] + DEBIT_CATEGORIES + ["Total Debit"] + NET_CATEGORIES + ["Total NET"] + BALANCE_ROWS[2:]

    if len(summary_df) != len(ordered_rows):
        print(summary_df.index.values)
        print("Not subset")
        44
        return

    summary_df = summary_df.reindex(ordered_rows)

    write_to_excel(year, month, transactions_dfs, summary_df)

if __name__ == "__main__":
    main()
